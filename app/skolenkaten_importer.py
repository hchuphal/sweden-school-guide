from __future__ import annotations

import json
import re
import unicodedata
from copy import deepcopy
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import requests
from openpyxl import load_workbook

SKOLENKATEN_SOURCE_PAGES: dict[int, str] = {
    2026: "https://www.skolinspektionen.se/skolenkaten/resultat-fran-skolenkaten/resultat-skolenkaten-2026/",
    2025: "https://www.skolinspektionen.se/skolenkaten/resultat-fran-skolenkaten/resultat-skolenkaten-2025/",
}
SKOLENKATEN_SOURCE_PAGE = SKOLENKATEN_SOURCE_PAGES[2026]
SKOLENKATEN_PUBLISHED_DATES = {2026: "2026-05-28", 2025: "2025-05-27"}

SKOLENKATEN_URLS: dict[int, dict[str, str]] = {
    2026: {
        "f0_guardians": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2026/skolenkaten-vardnadshavare-forskoleklass-2026.xlsx",
        "grade5_pupils": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2026/skolenkaten-elever-grundskola-ak-5-2026.xlsx",
        "grade8_pupils": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2026/skolenkaten-elever-grundskola-ak-8-2026.xlsx",
        "school_guardians": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2026/skolenkaten-vardnadshavare-grundskola-ak-1-9-2026.xlsx",
    },
    2025: {
        # The 2025 F0 filename uses "fklass", unlike the 2026 filename.
        "f0_guardians": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2025/skolenkaten-vardnadshavare-fklass-2025.xlsx",
        "grade5_pupils": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2025/skolenkaten-elever-grundskola-ak-5-2025.xlsx",
        "grade8_pupils": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2025/skolenkaten-elever-grundskola-ak-8-2025.xlsx",
        "school_guardians": "https://www.skolinspektionen.se/globalassets/02-beslut-rapporter-stat/statistik/statistik-skolenkaten/2025/skolenkaten-vardnadshavare-grundskola-ak-1-9-2025.xlsx",
    },
}

# Known Skolverket skolenhetskod values for schools that did not yet have a
# schoolUnitID URL in the original Göteborg seed data.
KNOWN_SCHOOL_UNIT_IDS: dict[str, int] = {
    "jattestensskolan": 55124975,
    "herrgardsskolan": 62479729,
    "taubeskolan": 11099088,
    "lerlyckeskolan": 48653651,
    "bjurslattsskolan": 37407044,
    "brackeskolan": 29466561,
    "innovitaskolan-st-jorgen": 27265374,
    "fridaskolan-kvillebacken": 56328727,
    "ebba-petterssons-privatskola": 88247874,
    "goteborgs-hogre-samskola-lilla-samskolan": 67568030,
    # ISGR has separate F-6 and 7-9 units. Use the F-6 unit for F0/F-6 survey data.
    "isgr": 93758282,
    "montessoriskolan-casa": 76938537,
    "montessoriskolan-centrum": 73995321,
}

SCHOOL_UNIT_RE = re.compile(r"schoolUnitID=(\d+)", re.IGNORECASE)

# Excel columns are 1-indexed, matching the published Skolenkäten files.
F0_COLUMNS = {
    "f0Satisfaction": 327,
    "safety": 168,
    "studyPeace": 141,
    "support": 68,
    "stimulation": 32,
}
PUPIL_COLUMNS = {
    "studentSatisfaction": 362,
    "pupilSafety": 264,
    "pupilStudyPeace": 237,
    "pupilSupport": 92,
    "pupilStimulation": 43,
}
GUARDIAN_COLUMNS = {
    "parentSatisfaction": 327,
    "guardianSafety": 168,
    "guardianStudyPeace": 141,
    "guardianSupport": 68,
    "guardianStimulation": 32,
}


def _normalise(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch)).lower()
    text = text.replace("–", "-").replace("—", "-")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _normalise_municipality(value: Any) -> str:
    text = _normalise(value)
    for suffix in (" kommun", " stad", " municipality"):
        if text.endswith(suffix):
            text = text[: -len(suffix)].strip()
    # Municipality labels often use the genitive form (Stockholms stad,
    # Göteborgs Stad, Mölndals kommun), while survey files use the base name.
    if text.endswith("s") and len(text) > 4:
        text = text[:-1]
    return text


def _normalise_school_name(value: Any) -> str:
    text = _normalise(value)
    replacements = {
        "grundskolan": "skola",
        "grundskola": "skola",
        "skolan": "skola",
        "forskoleklass": "",
    }
    tokens = [replacements.get(token, token) for token in text.split()]
    tokens = [token for token in tokens if token and not re.fullmatch(r"f?[0-9](?:-[0-9])?", token)]
    return " ".join(tokens)


def _to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        clean = value.strip().replace(",", ".")
        if clean in {"", "-", "*", "**", "***"}:
            return None
        value = clean
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return round(number, 1)


def _to_int(value: Any) -> int | None:
    if value is None or value == "-":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _read_metric_file(path: Path, metric_columns: dict[str, int]) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    by_id: dict[int, dict[str, Any]] = {}
    by_name: dict[tuple[str, str], list[dict[str, Any]]] = {}
    all_rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        school_unit_id = _to_int(row[3] if len(row) > 3 else None)
        school_name = row[4] if len(row) > 4 else None
        if not school_unit_id or not school_name:
            continue
        municipality = row[2] if len(row) > 2 else None
        item: dict[str, Any] = {
            "schoolUnitId": school_unit_id,
            "sourceSchoolName": str(school_name).strip(),
            "sourceMunicipality": str(municipality).strip() if municipality else None,
            "groupSize": _to_int(row[5] if len(row) > 5 else None),
            "responseCount": _to_int(row[6] if len(row) > 6 else None),
            "responseRate": _to_float(row[7] if len(row) > 7 else None),
        }
        for key, col in metric_columns.items():
            item[key] = _to_float(row[col - 1] if len(row) >= col else None)
        by_id[school_unit_id] = item
        key = (_normalise_municipality(municipality), _normalise_school_name(school_name))
        by_name.setdefault(key, []).append(item)
        all_rows.append(item)
    return {"byId": by_id, "byName": by_name, "rows": all_rows}


def _similarity(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if left == right:
        return 1.0
    left_tokens, right_tokens = set(left.split()), set(right.split())
    token_score = len(left_tokens & right_tokens) / max(1, len(left_tokens | right_tokens))
    sequence_score = SequenceMatcher(None, left, right).ratio()
    containment = 0.96 if left in right or right in left else 0.0
    return max(sequence_score, token_score, containment)


def _lookup_metric_row(
    index: dict[str, Any],
    school_unit_id: int | None,
    school_name: str | None,
    municipality: str | None,
    aliases: list[str] | None = None,
) -> tuple[dict[str, Any] | None, str | None]:
    if school_unit_id and school_unit_id in index["byId"]:
        return index["byId"][school_unit_id], "school-unit-id"

    municipality_key = _normalise_municipality(municipality)
    names = [school_name, *(aliases or [])]
    normalised_names = [_normalise_school_name(name) for name in names if name]
    for name_key in normalised_names:
        exact = index["byName"].get((municipality_key, name_key), [])
        if len(exact) == 1:
            return exact[0], "name-and-municipality"

    # Conservative fuzzy fallback: only within the same municipality and only
    # when one candidate is clearly better than the next.
    candidates: list[tuple[float, dict[str, Any]]] = []
    for row in index["rows"]:
        if municipality_key and _normalise_municipality(row.get("sourceMunicipality")) != municipality_key:
            continue
        source_name = _normalise_school_name(row.get("sourceSchoolName"))
        score = max((_similarity(name, source_name) for name in normalised_names), default=0.0)
        if score >= 0.88:
            candidates.append((score, row))
    candidates.sort(key=lambda item: item[0], reverse=True)
    if candidates and (len(candidates) == 1 or candidates[0][0] - candidates[1][0] >= 0.05):
        return candidates[0][1], "fuzzy-name-and-municipality"
    return None, None


def _download(url: str, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists() and path.stat().st_size > 0:
        return path
    response = requests.get(
        url,
        timeout=90,
        headers={
            "User-Agent": "SwedenSchoolGuide/0.18",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream;q=0.9,*/*;q=0.1",
        },
    )
    response.raise_for_status()
    content = response.content
    if not content.startswith(b"PK"):
        raise ValueError(f"Downloaded Skolenkäten source is not an XLSX file: {url}")
    path.write_bytes(content)
    return path


def ensure_source_files(year: int, cache_dir: Path, *, use_network: bool = True) -> dict[str, Path]:
    urls = SKOLENKATEN_URLS.get(year)
    if not urls:
        raise ValueError(f"No Skolenkäten source URLs configured for {year}")
    files: dict[str, Path] = {}
    for key, url in urls.items():
        filename = url.rsplit("/", 1)[-1]
        local_path = cache_dir / str(year) / filename
        if use_network:
            _download(url, local_path)
        elif not local_path.exists():
            raise FileNotFoundError(f"Missing cached source file: {local_path}")
        files[key] = local_path
    return files


def extract_school_unit_id(record: dict[str, Any]) -> int | None:
    if record.get("schoolUnitId"):
        return _to_int(record.get("schoolUnitId"))
    for source in record.get("sources", []) or []:
        match = SCHOOL_UNIT_RE.search(source.get("url", ""))
        if match:
            return int(match.group(1))
    return KNOWN_SCHOOL_UNIT_IDS.get(record.get("slug", ""))


def build_skolenkaten_import(
    baseline_records: list[dict[str, Any]],
    source_files: dict[str, Path],
    *,
    year: int,
    verified_label: str | None = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    f0 = _read_metric_file(source_files["f0_guardians"], F0_COLUMNS)
    grade5 = _read_metric_file(source_files["grade5_pupils"], PUPIL_COLUMNS)
    grade8 = _read_metric_file(source_files["grade8_pupils"], PUPIL_COLUMNS)
    guardians = _read_metric_file(source_files["school_guardians"], GUARDIAN_COLUMNS)

    updated: list[dict[str, Any]] = []
    matched_count = 0
    id_match_count = 0
    name_match_count = 0
    unmatched: list[str] = []
    source_page = SKOLENKATEN_SOURCE_PAGES[year]
    verified = verified_label or f"Skolenkäten {year}, imported from official Excel files"

    for source_record in baseline_records:
        record = deepcopy(source_record)
        record["dataYear"] = int(year)
        school_unit_id = extract_school_unit_id(record)
        municipality = record.get("municipality") or record.get("area")
        aliases = record.get("surveyAliases") or []

        f0_row, f0_method = _lookup_metric_row(f0, school_unit_id, record.get("name"), municipality, aliases)
        g5_row, g5_method = _lookup_metric_row(grade5, school_unit_id, record.get("name"), municipality, aliases)
        g8_row, g8_method = _lookup_metric_row(grade8, school_unit_id, record.get("name"), municipality, aliases)
        guardian_row, guardian_method = _lookup_metric_row(guardians, school_unit_id, record.get("name"), municipality, aliases)
        rows = [row for row in (f0_row, g5_row, g8_row, guardian_row) if row]
        methods = [method for method in (f0_method, g5_method, g8_method, guardian_method) if method]

        discovered_ids = {row.get("schoolUnitId") for row in rows if row.get("schoolUnitId")}
        if not school_unit_id and len(discovered_ids) == 1:
            school_unit_id = int(next(iter(discovered_ids)))
        record["schoolUnitId"] = school_unit_id
        has_match = bool(rows)

        if f0_row:
            for field in ["f0Satisfaction", "safety", "studyPeace", "support"]:
                record[field] = f0_row.get(field)

        pupil_row = g5_row or g8_row
        if pupil_row:
            record["studentSatisfaction"] = pupil_row.get("studentSatisfaction")
            if record.get("safety") is None:
                record["safety"] = pupil_row.get("pupilSafety")
            if record.get("studyPeace") is None:
                record["studyPeace"] = pupil_row.get("pupilStudyPeace")
            if record.get("support") is None:
                record["support"] = pupil_row.get("pupilSupport")

        if guardian_row:
            record["parentSatisfaction"] = guardian_row.get("parentSatisfaction")
            if record.get("safety") is None:
                record["safety"] = guardian_row.get("guardianSafety")
            if record.get("studyPeace") is None:
                record["studyPeace"] = guardian_row.get("guardianStudyPeace")
            if record.get("support") is None:
                record["support"] = guardian_row.get("guardianSupport")

        if has_match:
            matched_count += 1
            if any(method == "school-unit-id" for method in methods):
                id_match_count += 1
            else:
                name_match_count += 1
            record["lastVerified"] = SKOLENKATEN_PUBLISHED_DATES[year]
            existing_note = record.get("verificationNote") or ""
            method_label = "school-unit ID" if any(method == "school-unit-id" for method in methods) else "school name and municipality"
            record["verificationNote"] = (
                f"{verified}. Survey fields matched by {method_label}"
                f"{f' (skolenhetskod {school_unit_id})' if school_unit_id else ''}. "
                f"Academic and admission fields remain separate. {existing_note}"
            ).strip()
            sources = record.setdefault("sources", [])
            if not any(source_page in (src.get("url", "")) for src in sources):
                sources.append({"label": f"Skolinspektionen Skolenkäten {year}", "url": source_page})
            record["surveyImportStatus"] = "matched"
            record["surveyMatchMethods"] = sorted(set(methods))
            record["surveySourceDetails"] = {
                "schoolUnitId": school_unit_id,
                "f0Guardians": _source_summary(f0_row),
                "grade5Pupils": _source_summary(g5_row),
                "grade8Pupils": _source_summary(g8_row),
                "schoolGuardians": _source_summary(guardian_row),
            }
        else:
            record["surveyImportStatus"] = "not-found-in-skolenkaten-files"
            unmatched.append(record.get("name") or record.get("slug") or str(school_unit_id))

        updated.append(record)

    metadata = {
        "year": year,
        "sourcePage": source_page,
        "matchedSchools": matched_count,
        "matchedById": id_match_count,
        "matchedByNameAndMunicipality": name_match_count,
        "totalSchools": len(baseline_records),
        "unmatchedSchools": unmatched,
        "sourceFiles": {key: str(path) for key, path in source_files.items()},
        "note": "Survey fields are matched first by skolenhetskod and then conservatively by school name plus municipality. Missing or suppressed cells remain null. Academic and admission data are not changed by this importer.",
    }
    return updated, metadata


def _source_summary(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    return {
        "sourceSchoolName": row.get("sourceSchoolName"),
        "sourceMunicipality": row.get("sourceMunicipality"),
        "schoolUnitId": row.get("schoolUnitId"),
        "groupSize": row.get("groupSize"),
        "responseCount": row.get("responseCount"),
        "responseRate": row.get("responseRate"),
    }


def load_baseline(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if isinstance(payload, dict) and "schools" in payload:
        return payload["schools"]
    if isinstance(payload, list):
        return payload
    raise ValueError("Baseline must be a list or object with a schools array")


def write_import_json(records: list[dict[str, Any]], output_path: Path, metadata: dict[str, Any]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"metadata": metadata, "schools": records}
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return output_path
